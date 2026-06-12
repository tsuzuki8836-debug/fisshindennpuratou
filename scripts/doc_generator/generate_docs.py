import os
import json
import platform
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. 共通スタイル・環境定数定義
# ==========================================
FONT_NAME = "BIZ UDPゴシック"

FONT_TITLE = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
FONT_SECTION = Font(name=FONT_NAME, size=12, bold=True)
FONT_HEADER = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
FONT_REGULAR = Font(name=FONT_NAME, size=10, bold=False)

FILL_TITLE = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
FILL_HEADER = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
FILL_FAULT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

def auto_fit_columns(ws):
    """全角文字、改行、および余白を考慮した自動幅調整ロジック"""
    ws.views.sheetView[0].showGridLines = True
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                lines = str(cell.value).split('\n')
                for line in lines:
                    length = sum(2 if ord(c) > 128 else 1 for c in line)
                    if length > max_len:
                        max_len = length
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def draw_sheet_title(ws, title_text):
    """共通の大型上部シートヘッダー描画"""
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

# ✨ 【修正ポイント】引数リストの末尾に `is_test_sheet=False` を追加して、呼び出し側と完全一致させました
def render_table(ws, start_row, section_title, headers, rows, table_type="HORIZONTAL", is_test_sheet=False):
    """メタデータと描画ロジックを完全分離した、テンプレートマッピングエンジン"""
    ws.cell(row=start_row, column=1, value=section_title).font = FONT_SECTION
    
    header_row = start_row + 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 24
    
    current_row = header_row + 1
    for r_idx, row_data in enumerate(rows):
        ws.row_dimensions[current_row].height = 22 if table_type == "VERTICAL" else 20
        
        # デシジョンテーブルにおける異常系・境界値ケースの条件付きハイライト制御
        use_fault_fill = (is_test_sheet and row_data[0] in ["TC-PW-02", "TC-PW-03", "TC-PW-04"])
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = FONT_REGULAR
            cell.border = THIN_BORDER
            
            # テキスト長や数式、改行に応じた折り返しアラインメントの最適化
            if "Formula_" in str(val) or "<p>" in str(val) or "\n" in str(val) or len(str(val)) > 30:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center", horizontal="left" if table_type == "HORIZONTAL" and c_idx > 1 else "center")
                
            if use_fault_fill:
                cell.fill = FILL_FAULT
            elif table_type == "HORIZONTAL" and r_idx % 2 == 1:
                cell.fill = FILL_ZEBRA
            elif table_type == "VERTICAL" and c_idx == 1:
                cell.fill = FILL_ZEBRA  # 縦型テーブルは項目列をゼブラ色で識別
                
        current_row += 1
    return current_row + 1

# ==========================================
# 2. クロスプラットフォーム対応グラフィックエンジン
# ==========================================
def generate_mermaid_image():
    """OS判定・可観測性・Puppeteerクラッシュ対策を内包した画像化処理"""
    mermaid_text = """graph TD
    Start[トリガー: 依頼タスク回答の保存時<br/>状況 = '承認済み' かつ 更新時のみ] --> Dec_OwnerType{親タスクの所有者は<br/>User（ユーザー）か？}
    Dec_OwnerType -- YES --> Ass_UserAddresses[数式: Formula_ToAddresses<br/>所有者 + マネージャーのEmailを結合]
    Dec_OwnerType -- NO --> Ass_QueueAddresses[数式: Formula_ToAddresses<br/>キューの送信先Emailを抽出/Null安全]
    Ass_UserAddresses --> Ass_CombineCc[数式: Formula_CcAddresses<br/>CC1〜CC10を1ノードで一括結合]
    Ass_QueueAddresses --> Ass_CombineCc
    Ass_CombineCc --> Act_SendEmail[コアアクション:<br/>メールを送信]
    Act_SendEmail -->|正常完了| End_Success([フロー正常終了])
    Act_SendEmail -.->|障害発生| Act_CreateFaultLog[障害処理:<br/>エラーログレコードの作成]
    Act_CreateFaultLog --> End_Fault([フロー異常終了])
"""
    config_path = "puppeteerConfig.json"
    mmd_path = "flow.mmd"
    png_path = "flow.png"
    
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump({"args": ["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"]}, f)
    with open(mmd_path, "w", encoding="utf-8") as f:
        f.write(mermaid_text)
        
    cmd = "mmdc.cmd" if platform.system() == "Windows" else "mmdc"
    shell_opt = True if platform.system() == "Windows" else False
    
    try:
        subprocess.run([cmd, "-i", mmd_path, "-o", png_path, "-p", config_path], capture_output=True, text=True, shell=shell_opt, check=True)
        print("[SUCCESS] Mermaidロジックフロー図(PNG)を正常にレンダリングしました。")
    except (subprocess.CalledProcessError, FileNotFoundError) as e:
        stderr_msg = e.stderr if hasattr(e, 'stderr') else str(e)
        print(f"[INFO] ローカル環境のmmdc未検出、またはエラーのため画像埋め込みをスキップします。詳細:\n{stderr_msg}")
    finally:
        for path in [config_path, mmd_path]:
            if os.path.exists(path): os.remove(path)

# ==========================================
# 3. 設計書データセット定義 (テンプレート1対1追従)
# ==========================================
DATA_S1_T1 = {
    "section": "## 1. 基本情報",
    "headers": ["設定項目", "設定値（実際の値を記載）", "補足・仕様変更履歴"],
    "rows": [
        ["機能名", "依頼タスク回答：承認時完了メール通知フロー", "新規業務自動化ロジック策定"],
        ["API参照名", "RequestTaskAnswer_ApprovedNotificationFlow", "RequestTaskAnswer__cに対応するAfter-Saveフロー"],
        ["概要・目的", "依頼タスク回答の状況が「承認済み」になった瞬間、親タスクの所有者群(To)および設定されたCC1～CC10のユーザー群(Cc)へ対応完了をリアルタイム通知する。", "完了メールの自動配信化、最下部Chatterメンションは廃止"],
        ["対象オブジェクト", "依頼タスク回答 (RequestTaskAnswer__c)", "項目定義書シート16のカスタムオブジェクトに準拠"],
        ["トリガー（起動条件）", "更新された", "Status__cの変更イベントを保存後トリガーで捕捉するため"],
        ["実行タイミング", "保存後 (アクションと関連レコード)", "メール送信コアアクションおよび障害パスによるDML操作を実行するため"]
    ]
}

DATA_S1_T2 = {
    "section": "## 2. エントリ条件（プロパティ）",
    "headers": ["設定項目", "設定値（実際の値を記載）", "補足・仕様変更履歴"],
    "rows": [
        ["条件の要件", "すべての条件に一致 (AND)", "確実なトリガーフィルタ処理のため"],
        ["フローを実行するタイミング", "条件の要件に一致するようにレコードを更新したときのみ", "【ガードレール】他項目更新時の重複通知および再帰トリガーによる無限ループを完全排除"]
    ]
}

DATA_S1_T3 = {
    "section": "## 3. エントリ条件の詳細",
    "headers": ["項目ラベル", "API参照名", "演算子", "値 / 数式"],
    "rows": [
        ["状況", "Status__c", "次の文字列と一致する", "承認済み"]
    ]
}

DATA_S3_T1 = {
    "section": "## 1. リソース定義",
    "headers": ["リソース種別", "API参照名", "データ型", "コレクション", "説明・初期値・計算式"],
    "rows": [
        ["数式", "Formula_ToAddresses", "テキスト", "×", 
         "IF(\n    BEGINS({!$Record.RequestTask__r.OwnerId}, '005'),\n    {!$Record.RequestTask__r.Owner:User.Email} &\n    IF(NOT(ISBLANK({!$Record.RequestTask__r.Owner:User.ManagerId})), ',' & {!$Record.RequestTask__r.Owner:User.Manager.Email}, ''),\n    {!$Record.RequestTask__r.Owner:Group.Email}\n)\n\n[設計根拠: 親タスク所有者がキュー(00G)だった場合の実行時クラッシュを100%防ぐ防御的数式。]"],
        ["数式", "Formula_CcAddresses", "テキスト", "×", 
         "SUBSTITUTE(\n    IF(NOT(ISBLANK({!$Record.CC1__c})), ',' & {!$Record.CC1__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC2__c})), ',' & {!$Record.CC2__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC3__c})), ',' & {!$Record.CC3__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC4__c})), ',' & {!$Record.CC4__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC5__c})), ',' & {!$Record.CC5__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC6__c})), ',' & {!$Record.CC6__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC7__c})), ',' & {!$Record.CC7__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC8__c})), ',' & {!$Record.CC8__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC9__c})), ',' & {!$Record.CC9__r.Email}, '') &\n    IF(NOT(ISBLANK({!$Record.CC10__c})), ',' & {!$Record.CC10__r.Email}, ''),\n    ',', '', 1\n)\n\n[設計根拠: ループ要素を介さず1ノードでCC1〜CC10をカンマ結合。Bulkify制限を完全クリア。]"],
        ["テキストテンプレート", "tt_EmailBody", "テキスト", "×", 
         "<p>関係者各位</p><p>ご依頼のありましたタスクにつきまして、回答が承認され対応が完了いたしました。</p><hr><p>■ 依頼タスク内容</p><ul><li><b>親タスク名:</b> {!$Record.RequestTask__r.Name}</li><li><b>回答完了日:</b> {!$Flow.CurrentDate}</li></ul>\n\n[設計根拠: テンプレート指定に基づき、最下部のメンション文字列を完全に排除したHTML本文。]"]
    ]
}

DATA_S3_T2 = {
    "section": "## 2. 処理ロジック（ビジネスロジック）ステップ定義",
    "headers": ["ステップ名", "対象オブジェクト/コレクション", "条件（抽出/処理）", "処理内容/項目マッピング"],
    "rows": [
        ["コアアクション (メールを送信)", "N/A (標準アクション)", "エントリ条件の判定True時", 
         "• 送信本文 = {!tt_EmailBody}\n• 件名 = 【完了報告】依頼タスク回答が承認されました（タスク: {!$Record.RequestTask__r.Name}）\n• 受信者（To） = {!Formula_ToAddresses}\n• 受信者（Cc） = {!Formula_CcAddresses}\n• リッチテキスト本文として送信 = $GlobalConstant.True"],
        ["レコードを作成", "アプリケーションログ (ApplicationLog__c)", "【障害パス】メール送信エラー発生時", 
         "• エラーメッセージ = {!$Flow.FaultMessage}\n• 発生箇所 = RequestTaskAnswer_Flow\n\n[設計根拠: アドレス不正時の一括更新ロールバックを防ぎ、可観測性を担保。]"]
    ]
}

DATA_S3_T3 = {
    "section": "## 3. 分岐（決定）ロジック定義",
    "headers": ["決定要素名", "分岐名（条件名）", "分岐条件（判定ロジック）", "遷移先（次ステップ名）"],
    "rows": [
        ["Dec_OwnerType\n(親タスク所有者種別判定)", "User宛先生成", "項目: {!$Record.RequestTask__r.OwnerId}\n演算子: 次の文字列で始まる\n値: 005", "Ass_UserAddresses\n(所有者+マネージャーのEmail結合)"],
        ["Dec_OwnerType\n(親タスク所有者種別判定)", "デフォルト", "上記の条件に合致しない場合（キュー所有等）", "Ass_QueueAddresses\n(キューのEmailのみを抽出して割当)"]
    ]
}

DATA_S4_T1 = {
    "section": "## 1. テスト概要",
    "headers": ["検証プロパティ", "テスト概要定義（実際の定義を記載）"],
    "rows": [
        ["テスト項目名", "依頼タスク回答承認時の複数ルート（To/Cc）への完了通知メールリアルタイム配信およびNull安全性、大量データ一括処理（Bulkify）の検証"],
        ["テスト分類", "ガバナ制限/大量データ、ビジネスロジック、異常系エラーハンドリング"],
        ["前提条件 / データ", "親の依頼タスクがユーザーまたはキュー所有で存在すること。新規カスタム項目 CC1__c 〜 CC10__c がオブジェクトに追加されていること。"],
        ["該当の制限・観点", "送信エラー時の障害パス（Fault Path）動作、数式コンパイルサイズ（5,000バイト）制限の回避、データローダによる200件一括更新（Bulkification）時のCPU制限・SOQLクエリ制限のクリア観点。"]
    ]
}

DATA_S4_T2 = {
    "section": "## 2. デシジョンテーブル",
    "headers": ["テストケースID", "条件A（レコードの状況：Status__c）", "条件B（親タスクの所有者 / マネージャー設定）", "条件C（CC1〜CC10の入力パターン・状態）", "期待値（処理結果 / 更新内容 / エラーハンドリング）"],
    "rows": [
        ["TC-PW-01", "承認済み", "所有者:User(マネージャー設定あり)", "10項目すべてに有効なユーザーを設定", "Toに2名、Ccに10名のアドレスが正確にマッピングされメールが1通正常配信。200件一括更新でもガバナ制限をクリアすること。"],
        ["TC-PW-02", "承認済み", "所有者:キュー(Queue) (マネージャー概念なし)", "一部のみ入力あり(CC1, CC2のみ)", "Null Pointer Exceptionを起こさず、キューのグループEmail(To)およびCC1,CC2(Cc)宛てにエラーなくメール送信が完了すること。"],
        ["TC-PW-03", "承認済み", "所有者:User", "任意（設定されたアドレスのドメイン形式が不正）", "配信エラー時にフローがクラッシュせず、障害パス(Fault Path)が作動してApplicationLog__cレコードにエラー内容が正常に保存されること。"],
        ["TC-PW-04", "承認済み (変更なし)", "所有者:User", "任意", "エントリ条件の起動制限(Only when updated...)により、フロー自体が起動せず、不要な多重配信が100%防止されること。"],
        ["TC-PW-05", "未申請 / 却下", "任意", "任意", "トリガー条件に合致せず、フローは起動しないこと。"]
    ]
}

# ==========================================
# 5. メインExcelドキュメント自動生成
# ==========================================
def main():
    wb = Workbook()
    
    # --- シート1: 基本情報 ---
    ws1 = wb.active
    ws1.title = "基本情報"
    draw_sheet_title(ws1, "Salesforceフロー設計書（基本情報・エントリ条件）")
    next_row = render_table(ws1, 3, DATA_S1_T1["section"], DATA_S1_T1["headers"], DATA_S1_T1["rows"], table_type="VERTICAL")
    next_row = render_table(ws1, next_row, DATA_S1_T2["section"], DATA_S1_T2["headers"], DATA_S1_T2["rows"], table_type="VERTICAL")
    render_table(ws1, next_row, DATA_S1_T3["section"], DATA_S1_T3["headers"], DATA_S1_T3["rows"], table_type="HORIZONTAL")
    auto_fit_columns(ws1)
    
    # --- シート2: 処理フロー図 ---
    ws2 = wb.create_sheet(title="処理フロー図")
    ws2["A1"] = "## 2. 処理フロー図"
    ws2["A1"].font = FONT_SECTION
    generate_mermaid_image()
    if os.path.exists("flow.png"):
        from openpyxl.drawing.image import Image as OpenpyxlImage
        ws2.add_image(OpenpyxlImage("flow.png"), "A3")
    else:
        ws2["A3"] = "（Mermaid画像出力エンジン未検出のため、テキストフロー定義を設計として参照してください）"
        ws2["A3"].font = FONT_REGULAR
    ws2.views.sheetView[0].showGridLines = True
        
    # --- シート3: リソース定義・処理要素定義 ---
    ws3 = wb.create_sheet(title="リソース定義・処理要素定義")
    draw_sheet_title(ws3, "フロー構造詳細設計（統合リソース＆ロジック定義）")
    next_row = render_table(ws3, 3, DATA_S3_T1["section"], DATA_S3_T1["headers"], DATA_S3_T1["rows"], table_type="HORIZONTAL")
    next_row = render_table(ws3, next_row, DATA_S3_T2["section"], DATA_S3_T2["headers"], DATA_S3_T2["rows"], table_type="HORIZONTAL")
    render_table(ws3, next_row, DATA_S3_T3["section"], DATA_S3_T3["headers"], DATA_S3_T3["rows"], table_type="HORIZONTAL")
    auto_fit_columns(ws3)
    
    # --- シート4: テスト仕様書 ---
    ws4 = wb.create_sheet(title="テスト仕様書")
    draw_sheet_title(ws4, "高品質検証仕様書（デシジョンテーブルマトリクス）")
    next_row = render_table(ws4, 3, DATA_S4_T1["section"], DATA_S4_T1["headers"], DATA_S4_T1["rows"], table_type="VERTICAL")
    render_table(ws4, next_row, DATA_S4_T2["section"], DATA_S4_T2["headers"], DATA_S4_T2["rows"], table_type="HORIZONTAL", is_test_sheet=True)
    auto_fit_columns(ws4)
    
    # ファイル書き出し
    filename = "RequestTaskAnswerFlow_TemplateStrictDoc.xlsx"
    wb.save(filename)
    if os.path.exists("flow.png"): os.remove("flow.png")
    print(f"[SUCCESS] テンプレート定義に100%合致したクリーン設計書を生成しました: {filename}")

if __name__ == "__main__":
    main()